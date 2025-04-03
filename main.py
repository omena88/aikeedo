from fastapi import FastAPI, UploadFile, File, Form, Response
from fastapi.responses import FileResponse, StreamingResponse
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side
import os
import tempfile
import logging
import io
from fastapi.middleware.cors import CORSMiddleware
# Importar el router del módulo de margen
from margen import router as margen_router

app = FastAPI(title="Procesador de Cuentas por Cobrar",
             description="API para procesar archivos de cuentas por cobrar y generar reportes de Aging")

# Configuración CORS
# Lista de orígenes permitidos. Reemplaza "*" con tu dominio de WordPress en producción.
origins = [
    "*", # Permitir todos los orígenes (útil para desarrollo/pruebas)
    # "https://tu-dominio-wordpress.com", # Ejemplo para producción
    # "http://tu-dominio-wordpress.com", # Si también usa http
    # "http://localhost:xxxx" # Si pruebas desde un servidor local
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True, # Permitir cookies si las usas (importante para algunas configs)
    allow_methods=["POST", "OPTIONS"], # Métodos HTTP permitidos
    allow_headers=["*"], # Cabeceras permitidas (ej. Content-Type)
    expose_headers=["content-disposition"] # Cabeceras que el navegador puede leer (para el nombre del archivo)
)

# Configurar logging básico
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Incluir el router del módulo de margen
app.include_router(margen_router)

def procesar_cuentas_por_cobrar_excel(file_path, fecha_corte_str):
    logger.info(f"Iniciando procesamiento para archivo: {file_path}, fecha corte: {fecha_corte_str}")
    fecha_corte = datetime.strptime(fecha_corte_str, "%d/%m/%Y")
    with pd.ExcelFile(file_path) as xls:
        df = pd.read_excel(xls, sheet_name=xls.sheet_names[0], header=None)

    processed_data = []
    document_summary = {}
    current_account = None
    current_anexo = None
    current_client = None

    def format_date(value):
        date_value = pd.to_datetime(value, errors='coerce')
        return date_value.strftime("%d/%m/%Y") if pd.notna(date_value) else "Sin fecha"

    def convert_to_float(value):
        try:
            return round(float(value), 2)
        except:
            return 0.0

    for index, row in df.iterrows():
        if isinstance(row[1], str) and "Cuenta :" in row[1]:
            current_account = row[1].split(":")[1].strip().split()[0]

        elif isinstance(row[1], str) and "Anexo :" in row[1]:
            anexo_data = row[1].split(":")[1].strip().split(maxsplit=1)
            current_anexo = anexo_data[0][2:] if anexo_data[0].startswith("02") else anexo_data[0]
            current_client = anexo_data[1] if len(anexo_data) > 1 else ""

        elif isinstance(row[1], str) and "Saldo Anexo :" in row[1]:
            continue

        elif pd.notna(row[1]) and pd.notna(row[2]):
            compro = str(row[1]).strip()
            sec = str(row[2]).strip()
            fecha_vcto = format_date(row[6]) if pd.notna(row[6]) else "Sin fecha"
            documento = str(row[4]).strip()
            saldo_mn = convert_to_float(row[10]) - convert_to_float(row[11])
            saldo_me = convert_to_float(row[13]) - convert_to_float(row[14])

            if documento not in document_summary:
                document_summary[documento] = {
                    "Cuenta": current_account,
                    "Anexo": current_anexo,
                    "Cliente": current_client,
                    "Compro": compro,
                    "Sec": sec,
                    "Fecha Vcto": fecha_vcto,
                    "Saldo MN": 0.0,
                    "Saldo ME": 0.0
                }

            document_summary[documento]["Saldo MN"] += saldo_mn
            document_summary[documento]["Saldo ME"] += saldo_me

            if document_summary[documento]["Fecha Vcto"] == "Sin fecha":
                document_summary[documento]["Fecha Vcto"] = fecha_vcto

    aging_data = []
    aging_summary = {}
    for doc, data in document_summary.items():
        if round(data["Saldo MN"], 2) != 0:
            fecha_vcto_dt = datetime.strptime(data["Fecha Vcto"], "%d/%m/%Y") if data["Fecha Vcto"] != "Sin fecha" else None
            dias_vencidos = (fecha_corte - fecha_vcto_dt).days if fecha_vcto_dt else 0

            categoria = (
                "VIGENTE" if dias_vencidos <= 0 else
                "001-30 días" if dias_vencidos <= 30 else
                "031-60 días" if dias_vencidos <= 60 else
                "061-90 días" if dias_vencidos <= 90 else
                "091-120 días" if dias_vencidos <= 120 else
                "121-150 días" if dias_vencidos <= 150 else
                "151-180 días" if dias_vencidos <= 180 else
                "181-210 días" if dias_vencidos <= 210 else
                "211-240 días" if dias_vencidos <= 240 else
                "241-270 días" if dias_vencidos <= 270 else
                "271-300 días" if dias_vencidos <= 300 else
                "301-330 días" if dias_vencidos <= 330 else
                "331-360 días" if dias_vencidos <= 360 else
                "361 a más"
            )

            aging_data.append([
                data["Cuenta"], data["Anexo"], data["Cliente"], doc,
                data["Compro"], data["Sec"], data["Fecha Vcto"],
                dias_vencidos, round(data["Saldo MN"], 2), categoria
            ])

            aging_summary[categoria] = aging_summary.get(categoria, 0.0) + data["Saldo MN"]

    df_aging = pd.DataFrame(aging_data, columns=[
        "Cuenta", "Anexo", "Cliente", "Documento", "Compro", "Sec",
        "Fecha Vcto", "Días Vencidos", "Saldo MN", "Aging"
    ])

    total_saldo = sum(aging_summary.values())
    resumen_rows = [["Categoría", "Saldo MN", "%"]]
    for cat in [
        "VIGENTE", "001-30 días", "031-60 días", "061-90 días",
        "091-120 días", "121-150 días", "151-180 días", "181-210 días",
        "211-240 días", "241-270 días", "271-300 días", "301-330 días",
        "331-360 días", "361 a más"
    ]:
        saldo = round(aging_summary.get(cat, 0.0), 2)
        porcentaje = round((saldo / total_saldo) * 100, 2) if total_saldo != 0 else 0
        resumen_rows.append([cat, saldo, f"{porcentaje}%"])
    resumen_rows.append(["TOTAL", round(total_saldo, 2), "100%"])

    # Crear archivo Excel
    wb = Workbook()
    ws_aging = wb.active
    ws_aging.title = "Aging"
    for r in dataframe_to_rows(df_aging, index=False, header=True):
        ws_aging.append(r)

    ws_resumen = wb.create_sheet("Resumen")
    ws_resumen.append(["Resumen de Cuentas por Cobrar"])
    ws_resumen.append([f"Fecha de corte: {fecha_corte.strftime('%d/%m/%Y')}"])
    ws_resumen.append([])

    for row in resumen_rows:
        ws_resumen.append(row)

    # Estilo en la hoja "Resumen"
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws_resumen.iter_rows(min_row=5, max_row=ws_resumen.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = border_style
            cell.alignment = center_align
        row[0].font = bold_font

    ws_resumen["A1"].font = Font(bold=True, size=12)
    ws_resumen["A2"].font = Font(italic=True)

    # Guardar en un archivo temporal
    temp_dir = tempfile.gettempdir()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S_%f") # Añadido microsegundos para más unicidad
    output_path = os.path.join(temp_dir, f"cuentas_por_cobrar_{timestamp}.xlsx")
    logger.info(f"Intentando guardar archivo Excel en: {output_path}")
    wb.save(output_path)
    logger.info(f"Archivo Excel guardado exitosamente en: {output_path}")
    return output_path

@app.post("/procesar-cuentas-por-cobrar")
async def procesar_cuentas_por_cobrar(
    file: UploadFile = File(...),
    fecha_corte: str = Form(...)
):
    logger.info(f"Recibida solicitud para procesar archivo: {file.filename}, fecha_corte: {fecha_corte}")
    temp_file_path = None
    output_path = None
    output_bytes = None
    try:
        # Guardar el archivo temporalmente
        temp_dir = tempfile.gettempdir()
        # Usar un nombre de archivo temporal más robusto
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=temp_dir) as temp_file:
            content = await file.read()
            temp_file.write(content)
            temp_file_path = temp_file.name
        logger.info(f"Archivo temporal guardado en: {temp_file_path}")

        # Procesar el archivo para obtener la ruta del resultado
        logger.info("Llamando a procesar_cuentas_por_cobrar_excel...")
        output_path = procesar_cuentas_por_cobrar_excel(temp_file_path, fecha_corte)
        logger.info(f"Procesamiento completado. Archivo resultante en disco: {output_path}")

        # Leer el contenido del archivo generado en memoria
        try:
            with open(output_path, 'rb') as f:
                output_bytes = f.read()
            logger.info(f"Archivo leído en memoria, tamaño: {len(output_bytes)} bytes")
        except Exception as e_read:
            logger.error(f"Error al leer el archivo generado {output_path} en memoria: {e_read}", exc_info=True)
            raise

        # Devolver el contenido directamente
        file_name_for_download = os.path.basename(output_path)
        logger.info(f"Preparando Response con contenido binario para: {file_name_for_download}")
        return Response(
            content=output_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={f'Content-Disposition': f'attachment; filename="{file_name_for_download}"'} # Cabecera para descarga
        )

    except Exception as e:
        logger.error(f"Error en el endpoint /procesar-cuentas-por-cobrar: {e}", exc_info=True)
        return {"error": f"Ocurrió un error procesando el archivo: {str(e)}"}

    finally:
        # Asegurar limpieza de archivos temporales (subido y generado)
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
                logger.info(f"Archivo temporal de subida eliminado: {temp_file_path}")
            except Exception as e_remove_temp:
                logger.error(f"Error al eliminar archivo temporal de subida {temp_file_path}: {e_remove_temp}")
        if output_path and os.path.exists(output_path):
            try:
                os.remove(output_path)
                logger.info(f"Archivo temporal generado eliminado: {output_path}")
            except Exception as e_remove_output:
                logger.error(f"Error al eliminar archivo temporal generado {output_path}: {e_remove_output}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000) 
